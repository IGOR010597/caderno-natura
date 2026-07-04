from __future__ import annotations

import re
import json
import logging
import os
import base64
import zlib
from collections import OrderedDict
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from PIL import Image, ImageEnhance, ImageFilter, ImageOps


logger = logging.getLogger(__name__)


QUANTITY_MARKER = r"(?:[-–—:*=]|[xX]|qtd\.?|quantidade)?"
LINE_PATTERN = re.compile(
    rf"^\s*([A-Za-z0-9]+)\s*{QUANTITY_MARKER}\s*(\d+)?\s*$",
    re.IGNORECASE,
)


def parse_ocr_text(text: str) -> list[dict]:
    """Transform every non-empty OCR line into a reviewable row."""
    rows: list[dict] = []
    for raw in (line.strip() for line in text.splitlines()):
        if not raw:
            continue
        match = LINE_PATTERN.match(raw)
        if not match:
            rows.append({"code": "", "quantity": None, "status": "Revisar", "raw_line": raw})
            continue

        code, quantity_text = match.groups()
        quantity = int(quantity_text) if quantity_text else None
        valid = code.isdigit() and quantity is not None and quantity > 0
        rows.append(
            {
                "code": code,
                "quantity": quantity,
                "status": "OK" if valid else "Revisar",
                "raw_line": raw,
            }
        )
    return rows


AI_PROMPT = """
Analise esta foto de uma página de caderno com um pedido de produtos Natura.
Extraia SOMENTE o código numérico do produto e sua quantidade inteira em cada linha.
O primeiro número da linha é o código; o segundo é a quantidade. Considere formatos
como '123456 2', '123456 - 2', '123456 x2', '123456 qtd 2' e '123456 quantidade 2'.
Preserve zeros à esquerda. Não inclua nome, cliente, preço, valor ou observações.
Não invente dígitos escondidos ou ilegíveis. Marque confiança como 'baixa' se houver
qualquer dúvida em um dígito ou se código/quantidade estiverem incompletos. Inclua em
source_text a anotação curta que originou cada item. Retorne os itens na ordem da foto.
"""

AI_SCHEMA = {
    "type": "object",
    "properties": {
        "items": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "code": {"type": "string", "description": "Código exatamente como visto."},
                    "quantity": {"type": ["integer", "null"]},
                    "confidence": {"type": "string", "enum": ["alta", "baixa"]},
                    "source_text": {"type": "string"},
                },
                "required": ["code", "quantity", "confidence", "source_text"],
            },
        }
    },
    "required": ["items"],
}


def normalize_ai_items(items: list[dict]) -> list[dict]:
    rows: list[dict] = []
    for item in items:
        code = str(item.get("code", "")).strip().replace(" ", "")
        quantity = item.get("quantity")
        if isinstance(quantity, bool) or not isinstance(quantity, int):
            quantity = None
        confident = item.get("confidence") == "alta"
        valid = code.isdigit() and quantity is not None and quantity > 0 and confident
        rows.append({
            "code": code,
            "quantity": quantity,
            "status": "OK" if valid else "Revisar",
            "raw_line": str(item.get("source_text", "")).strip(),
        })
    return rows


def normalize_image(content: bytes, max_dimension: int = 1800) -> bytes:
    """Convert mobile formats, apply EXIF rotation and cap upload size for the AI."""
    try:
        from pillow_heif import register_heif_opener

        register_heif_opener()
        image = Image.open(BytesIO(content))
        image = ImageOps.exif_transpose(image)
        image.thumbnail((max_dimension, max_dimension), Image.Resampling.LANCZOS)
        if image.mode != "RGB":
            background = Image.new("RGB", image.size, "white")
            if "A" in image.getbands():
                background.paste(image, mask=image.getchannel("A"))
            else:
                background.paste(image)
            image = background
        output = BytesIO()
        image.save(output, format="JPEG", quality=86, optimize=True)
        return output.getvalue()
    except Exception as exc:
        raise ValueError("A foto não pôde ser convertida. Tente enviá-la como JPG.") from exc


def extract_rows_with_gemini(content: bytes, mime_type: str) -> list[dict]:
    api_key = os.getenv("GEMINI_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError("GEMINI_API_KEY não configurada.")

    from google import genai
    from google.genai import types

    client = genai.Client(api_key=api_key)
    primary = os.getenv("GEMINI_MODEL", "gemini-2.5-flash-lite")
    fallback = os.getenv("GEMINI_FALLBACK_MODEL", "gemini-2.5-flash")
    models = list(dict.fromkeys([primary, fallback]))
    last_error: Exception | None = None

    for model in models:
        try:
            response = client.models.generate_content(
                model=model,
                contents=[types.Part.from_bytes(data=content, mime_type=mime_type), AI_PROMPT],
                config=types.GenerateContentConfig(
                    response_mime_type="application/json",
                    response_json_schema=AI_SCHEMA,
                    temperature=0,
                ),
            )
            payload = json.loads(response.text or "{}")
            return normalize_ai_items(payload.get("items", []))
        except Exception as exc:
            last_error = exc
            logger.warning("Gemini model %s failed: %s: %s", model, type(exc).__name__, exc)
            error_text = str(exc).upper()
            if any(marker in error_text for marker in ("401", "403", "API_KEY_INVALID", "PERMISSION_DENIED")):
                break

    assert last_error is not None
    error_text = str(last_error).upper()
    status_code = getattr(last_error, "status_code", None) or getattr(last_error, "code", None)
    if status_code == 429 or "429" in error_text or "RESOURCE_EXHAUSTED" in error_text:
        message = "O limite gratuito da IA foi atingido. Tente novamente em um minuto."
    elif status_code in (401, 403) or "API_KEY_INVALID" in error_text or "PERMISSION_DENIED" in error_text:
        message = "A chave da IA não foi autorizada."
    elif status_code == 400 or "INVALID_ARGUMENT" in error_text:
        message = "A IA não aceitou o conteúdo desta foto."
    elif isinstance(last_error, (json.JSONDecodeError, KeyError, TypeError, ValueError)):
        message = "A IA retornou uma leitura inválida."
    else:
        message = "A leitura com IA está temporariamente indisponível."
    raise RuntimeError(message) from last_error


def extract_text_from_image(content: bytes) -> str:
    try:
        import pytesseract
    except ImportError as exc:  # pragma: no cover - deployment guard
        raise RuntimeError("O mecanismo de OCR não está instalado no servidor.") from exc

    try:
        image = Image.open(BytesIO(content))
        image.verify()
        image = Image.open(BytesIO(content)).convert("L")
    except Exception as exc:
        raise ValueError("A imagem enviada não pôde ser lida.") from exc

    # Upscaling and contrast help Tesseract with photographed notebook pages.
    if image.width < 1800:
        ratio = 1800 / image.width
        image = image.resize((1800, int(image.height * ratio)))
    image = ImageOps.autocontrast(image)
    image = ImageEnhance.Contrast(image).enhance(1.6)
    image = image.filter(ImageFilter.SHARPEN)

    try:
        return pytesseract.image_to_string(image, lang="por", config="--psm 6")
    except pytesseract.TesseractNotFoundError as exc:
        raise RuntimeError(
            "Tesseract OCR não encontrado. Instale o Tesseract e reinicie o sistema."
        ) from exc
    except pytesseract.TesseractError:
        # Some installations do not include Portuguese; numbers work with English.
        try:
            return pytesseract.image_to_string(image, lang="eng", config="--psm 6")
        except Exception as exc:
            raise RuntimeError("Não foi possível processar a imagem com o OCR.") from exc


def aggregate_products(products: list[dict]) -> OrderedDict[str, int]:
    aggregated: OrderedDict[str, int] = OrderedDict()
    for item in products:
        code = str(item.get("code", "")).strip()
        quantity = item.get("quantity")
        if not code:
            raise ValueError("Há um produto com código vazio.")
        if not code.isdigit():
            raise ValueError(f"O código '{code}' deve conter apenas números.")
        if isinstance(quantity, bool) or not isinstance(quantity, int) or quantity <= 0:
            raise ValueError(f"A quantidade do código {code} deve ser um inteiro maior que zero.")
        aggregated[code] = aggregated.get(code, 0) + quantity
    if not aggregated:
        raise ValueError("Adicione ao menos um produto antes de gerar a planilha.")
    return aggregated


def create_workbook(products: OrderedDict[str, int], destination: Path) -> None:
    workbook = build_workbook(products)
    workbook.save(destination)


def build_workbook(products: OrderedDict[str, int]) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["CÓDIGO", "QT"])
    for code, quantity in products.items():
        sheet.append([code, quantity])
        sheet.cell(sheet.max_row, 1).number_format = "@"
    sheet["A1"].font = Font(bold=True)
    sheet["B1"].font = Font(bold=True)
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 10
    return workbook


def create_workbook_bytes(products: OrderedDict[str, int]) -> bytes:
    output = BytesIO()
    build_workbook(products).save(output)
    return output.getvalue()


def encode_share_token(products: OrderedDict[str, int]) -> str:
    payload = json.dumps({"p": list(products.items())}, separators=(",", ":"))
    compressed = zlib.compress(payload.encode("utf-8"), level=9)
    return base64.urlsafe_b64encode(compressed).rstrip(b"=").decode("ascii")


def decode_share_token(token: str) -> OrderedDict[str, int]:
    if not token or len(token) > 12_000:
        raise ValueError("Link de compartilhamento inválido.")
    try:
        padded = token + "=" * (-len(token) % 4)
        compressed = base64.b64decode(padded, altchars=b"-_", validate=True)
        decompressor = zlib.decompressobj()
        raw = decompressor.decompress(compressed, 100_001)
        if len(raw) > 100_000 or decompressor.unconsumed_tail:
            raise ValueError
        payload = json.loads(raw.decode("utf-8"))
        items = payload["p"]
        if not isinstance(items, list) or len(items) > 500:
            raise ValueError
        return aggregate_products([
            {"code": item[0], "quantity": item[1]}
            for item in items
            if isinstance(item, list) and len(item) == 2
        ])
    except Exception as exc:
        raise ValueError("Link de compartilhamento inválido.") from exc
