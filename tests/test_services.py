from collections import OrderedDict

import pytest
from openpyxl import load_workbook

from app.services import aggregate_products, create_workbook, normalize_ai_items, parse_ocr_text


def test_parse_supported_notations():
    rows = parse_ocr_text(
        "123456 2\n234567 - 3\n345678 x2\n456789 qtd 4\n567890 quantidade 5"
    )
    assert [(row["code"], row["quantity"], row["status"]) for row in rows] == [
        ("123456", 2, "OK"),
        ("234567", 3, "OK"),
        ("345678", 2, "OK"),
        ("456789", 4, "OK"),
        ("567890", 5, "OK"),
    ]


def test_parse_preserves_unrecognized_lines_for_review():
    rows = parse_ocr_text("12A456 2\ntexto impossível\n123456")
    assert all(row["status"] == "Revisar" for row in rows)
    assert rows[1]["raw_line"] == "texto impossível"
    assert rows[2]["quantity"] is None


def test_normalize_ai_items_marks_low_confidence_for_review():
    rows = normalize_ai_items([
        {"code": "001234", "quantity": 2, "confidence": "alta", "source_text": "001234 x2"},
        {"code": "78910?", "quantity": 1, "confidence": "baixa", "source_text": "78910? 1"},
    ])
    assert rows[0] == {"code": "001234", "quantity": 2, "status": "OK", "raw_line": "001234 x2"}
    assert rows[1]["status"] == "Revisar"


def test_aggregate_sums_duplicates_and_keeps_code_as_string():
    result = aggregate_products([
        {"code": "001234", "quantity": 2},
        {"code": "789101", "quantity": 1},
        {"code": "001234", "quantity": 3},
    ])
    assert result == OrderedDict([("001234", 5), ("789101", 1)])


@pytest.mark.parametrize("product", [
    {"code": "", "quantity": 1},
    {"code": "ABC", "quantity": 1},
    {"code": "123", "quantity": 0},
    {"code": "123", "quantity": -1},
])
def test_aggregate_rejects_invalid_products(product):
    with pytest.raises(ValueError):
        aggregate_products([product])


def test_workbook_has_exact_natura_structure(tmp_path):
    destination = tmp_path / "pedido_natura.xlsx"
    create_workbook(OrderedDict([("001234", 5), ("789101", 1)]), destination)
    workbook = load_workbook(destination)
    assert workbook.sheetnames == ["Sheet1"]
    sheet = workbook["Sheet1"]
    assert sheet.max_row == 3
    assert sheet.max_column == 2
    assert sheet["A1"].value == "CÓDIGO"
    assert sheet["B1"].value == "QT"
    assert sheet["A2"].value == "001234"
    assert sheet["A2"].number_format == "@"
    assert sheet["B2"].value == 5
    assert isinstance(sheet["B2"].value, int)
